import openpyxl
import csv
import os
compareItemTrackingList = []
new_list = []
compareColumn = input('Please enter the name of the Column which will be used to compare all the data. This Column must be in all sheets.' '\n')
rowHeaderList = [compareColumn]
folderPath = input('Please enter the full folder path of the folder containing the spread sheets like the example listed. ' r'C:\users\your username\folder with spreadsheets' '\n')
exportFile = ''
exportChoice1 = input('Do you wish to have the exported file saved in the same folder? ' 'Reply Yes or No' '\n')
if exportChoice1.upper() == 'YES' or exportChoice1.upper() == 'Y':
    exportFile = folderPath
else:
    exportFile = input('Please enter the full folder path where you want the file saved. Enter the path in the same format as the example provided. ' r'C:\users\your username\documents' '\n')
exportFileNameChoice = input('Please choose the name of the export file. If you wish it to be a CSV or Excel file include it in the name. Your file name should look like example.csv or example.xlsx' '\n' )
exportFile = exportFile + '\\' + exportFileNameChoice
files = os.listdir(folderPath)

#Cleans _ and leading/trailing spaces off items passed into it
def cleanLineItem(lineItem):
    itemreturn = str(lineItem)
    itemreturn = itemreturn.strip()
    return itemreturn

def cleanLineItemHeader(lineItem):
    itemreturn = str(lineItem)
    itemreturn = itemreturn.replace('_', ' ')
    itemreturn = itemreturn.strip()
    itemreturn = itemreturn.title()
    return itemreturn

def cleanDictLine(linedict):
    temp_dict = {}
    for item in linedict:
        key = item[0]
        value = item[1]
        keyClean = cleanLineItemHeader(key)
        valueClean = value.strip()
        temp_dict.update({keyClean: valueClean})
    return temp_dict

def headerConstuctorCsv(csvfile, headerVeriableStore):
    with open(csvfile, 'r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for line in csv_reader:
            for key in line.keys():
                cleanKey = cleanLineItemHeader(key)
                if cleanKey not in headerVeriableStore and cleanKey != "":
                    headerVeriableStore.append(cleanKey)

def pullRecordssCsv(csvdoc,compareItemTrackingList,new_list,rowHeaders):
    with open(csvdoc, 'r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for line in csv_reader:
            clean_line = cleanDictLine(line.items())
            record = clean_line.get(compareColumn)
            if record not in compareItemTrackingList:
                compareItemTrackingList.append(record)
                new_list.append(clean_line)
            if record in compareItemTrackingList:
                index = compareItemTrackingList.index(record)
                #if the evaluated record has data but the new list does not have this yet the if statment bellow will add it to the end output data, if the row header does on exist in this data set it will skip the item in the record and check the next columb
                for item in rowHeaders:
                    try:
                        if new_list[index][item] == '' and clean_line[item] != '':
                            new_list[index][item] = clean_line[item]
                    except KeyError:
                        pass

def pullRecordsExcel(excelDoc,compareItemTrackingList,new_list,rowHeaders):
    workBook = openpyxl.load_workbook(excelDoc)
    excelSheet = workBook.active
    temp_store = tuple(excelSheet.values)
    headerRow = True
    temp_headers = []
    for row in temp_store:
        if headerRow == True:
            for item in row:
                cleanItem = cleanLineItemHeader(item)
                temp_headers.append(cleanItem)
                if cleanItem not in rowHeaders:
                    rowHeaders.append(cleanItem)
            headerRow = False
        else:
            index = 0
            temp_dict = {}
            for item in row:
                temp_dict.update({temp_headers[index]: item})
                index += 1
            newListIndex = ""
            try:
                newListIndex = compareItemTrackingList.index(temp_dict.get(compareColumn))
            except ValueError:
                apend_dict = {}
                compareItemTrackingList.append(temp_dict.get(compareColumn))
                for header in rowHeaderList:
                    apend_dict.update({header: temp_dict.get(header, '')})
                new_list.append(apend_dict)
                newListIndex = compareItemTrackingList.index(temp_dict.get(compareColumn))
            for header in rowHeaderList:
                try:
                    if new_list[newListIndex][header] == '' and temp_dict[header] != '':
                        new_list[index][item] = temp_dict[header]
                except KeyError:
                    pass
    workBook.close()

def csvExporter(outputfilename,headernames):
    with open(outputfilename, 'w', newline= "") as new_file:
        csv_writer = csv.DictWriter(new_file, headernames)
        csv_writer.writeheader()
        for item in new_list:
            csv_writer.writerow(item)

def excelExporter (excelDoc=exportFile,new_list=new_list,rowHeaders=rowHeaderList):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    row = 1
    headerRow = True
    column = 1
    if headerRow == True:
        for item in rowHeaders:
            ws1.cell(row=row, column=column, value=item)
            column += 1
        headerRow = False
        row += 1
        column = 1
    for eachRow in new_list:
        for value in eachRow.values():
            ws1.cell(row=row, column=column, value=value)
            column += 1
        row += 1
        column = 1
    wb.save(excelDoc)      


for file in files:
    if '.csv' in file or '.CSV' in file:
       pass
       fullfilepath = folderPath + '\\' + file
       headerConstuctorCsv(fullfilepath,rowHeaderList)
       pullRecordssCsv(fullfilepath,compareItemTrackingList,new_list,rowHeaderList)
    if '.xl' in file or '.XL' in file:
        fullfilepath = folderPath + '\\' + file
        pullRecordsExcel(fullfilepath,compareItemTrackingList,new_list,rowHeaderList)

if '.csv' in exportFileNameChoice:
    csvExporter(exportFile,rowHeaderList)
elif '.xlsx' in exportFileNameChoice:
    excelExporter()
else:
    print('The export file type must be an Excel xlsx or a csv document')